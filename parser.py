"""교육과정 학점 배당표 엑셀 파서 — BytesIO/경로 모두 지원"""
import warnings
warnings.filterwarnings("ignore")

import openpyxl
import re
import unicodedata
from pathlib import Path
from io import BytesIO


def _norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = s.replace("\n", " ").replace("\u3000", " ")
    return re.sub(r"\s+", " ", s).strip()


def _norm_name(s):
    """\uacfc\ubaa9\uba85 \uc804\uc6a9 \uc815\uaddc\ud654. NFKC\ub294 \ub85c\ub9c8\uc22b\uc790 \u2160/\u2161(U+2160..)\ub97c \uc601\ubb38 I/II\ub85c \ubc14\uafd4\ubc84\ub824
    \ud45c\uae30 \uac80\uc0ac(E1)\u00b7\uc704\uacc4 \uac80\uc0ac(C2)\ub97c \ub9dd\uac00\ub728\ub9ac\ubbc0\ub85c, \ub85c\ub9c8\uc22b\uc790\ub97c \ubcf4\uc874\ud558\ub294 NFC\ub9cc \uc801\uc6a9."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s))
    s = s.replace("\n", " ").replace("\u3000", " ")
    return re.sub(r"\s+", " ", s).strip()


def _to_num(v):
    """'28~30', '[2]', '3(택2)' 등에서 첫 숫자 추출"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"\d+(?:\.\d+)?", str(v))
    return float(m.group()) if m else None


def _norm_subject_type(t):
    """과목유형 표기 정규화.

    학교별로 '일반/진로/융합'처럼 약어를 쓰지만, 자율점검표(checker)는
    표준어 '일반선택/진로선택/융합선택'을 기대한다. 부분일치로 흡수해
    '일반', '일반선택', '전공 일반' 등을 모두 같은 표준어로 매핑한다.
    """
    t = _norm(t)
    key = t.replace(" ", "")
    if "공통" in key:
        return "공통"
    if "진로" in key:
        return "진로선택"
    if "융합" in key:
        return "융합선택"
    if "일반" in key:
        return "일반선택"
    return t


def find_target_sheet(wb):
    """'xxxx입학생' 시트 탐지 (괄호 부기·공백 허용)"""
    for name in wb.sheetnames:
        if re.search(r"\d{4}\s*입학생", name):
            return name
    return None


def official_subjects(wb):
    """숨김 '데이터베이스' 시트의 평면표 '과목명' 열에서 정규(표준) 과목명 집합 추출.
    이게 있으면 배당표 과목 중 목록에 없는 = 비표준(고시외/교육감승인/특목/오타) 과목을
    걸러낼 수 있다. 시트/헤더가 없으면 빈 set 반환(검증 생략)."""
    sheet = next((s for s in wb.sheetnames if "데이터베이스" in s.replace(" ", "")), None)
    if not sheet:
        return set()
    ws = wb[sheet]
    hdr_r = hdr_c = None
    for r in range(1, min(ws.max_row, 40) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            if _norm(ws.cell(r, c).value) == "과목명":
                hdr_r, hdr_c = r, c
                break
        if hdr_r:
            break
    if not hdr_r:
        return set()
    out = set()
    for r in range(hdr_r + 1, ws.max_row + 1):
        v = ws.cell(r, hdr_c).value
        if v not in (None, ""):
            out.add(_norm_name(v))
    return out


def find_data_range(cell, max_row):
    """헤더/데이터/합계 영역 행 번호 자동 탐지 (병합 인식 cell 사용)"""
    header_row, summary_start = None, None
    for r in range(1, min(max_row, 30) + 1):
        if _norm(cell(r, 2)) == "교과(군)":
            header_row = r
            break
    for r in range(max_row, 0, -1):
        for c in range(1, 5):
            v = _norm(cell(r, c)).replace(" ", "")
            if "교과이수학점소계" in v:
                summary_start = r
                break
        if summary_start:
            break
    # data_start: 헤더~합계 사이에서 '과목명이 있고 학점이 숫자인' 첫 행을 스캔.
    # 헤더 높이·spacer 행이 파일마다 달라 고정 오프셋(header+4)은 부정확하다.
    # (실제로 어떤 학교 파일에서는 첫 과목이 통째로 누락됐다)
    data_start = None
    if header_row and summary_start:
        for r in range(header_row + 1, summary_start):
            subj = _norm(cell(r, 4))
            base, run = _to_num(cell(r, 5)), _to_num(cell(r, 6))
            if subj and (base is not None or run is not None):
                data_start = r
                break
    return header_row, data_start, summary_start


def parse_curriculum(source):
    """
    source: 파일 경로(str/Path) 또는 BytesIO 객체 모두 허용
    """
    # ----- 워크북 로드 (타입에 따라 분기) -----
    wb = openpyxl.load_workbook(source, data_only=True)

    정규과목 = official_subjects(wb)  # 데이터베이스 시트의 정규 과목명 집합(없으면 빈 set)

    sheet = find_target_sheet(wb)
    if not sheet:
        raise ValueError(
            f"'xxxx입학생' 시트를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}"
        )
    ws = wb[sheet]

    # ----- 병합 셀 forward-fill (탐지보다 먼저: find_data_range가 병합 인식) -----
    # vmerged: 세로(여러 행) 병합에 속한 칸. 선택그룹 학기칸(예: "12(택4)")이 행마다
    # 복제되는데, 이건 과목별 배치가 아니라 그룹 총합이라 배치합 계산에서 제외해야 한다.
    mmap = {}
    vmerged = set()
    for mr in ws.merged_cells.ranges:
        top = ws.cell(mr.min_row, mr.min_col).value
        is_vert = mr.max_row > mr.min_row
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                mmap[(r, c)] = top
                if is_vert:
                    vmerged.add((r, c))

    def cell(r, c):
        return mmap.get((r, c), ws.cell(r, c).value)

    h, ds, ss = find_data_range(cell, ws.max_row)
    if not (h and ds and ss):
        raise ValueError(
            f"데이터 범위 탐지 실패 (header={h}, data_start={ds}, summary_start={ss})"
        )

    # ----- 과목 행 파싱 -----
    rows = []
    last_g, last_t = "", ""
    for r in range(ds, ss):
        gun = _norm(cell(r, 2))
        typ = _norm(cell(r, 3))
        subj = _norm_name(cell(r, 4))
        if gun:
            last_g = gun
        else:
            gun = last_g
        if typ:
            last_t = typ
        else:
            typ = last_t
        if not subj:
            continue
        # 교차이수 [ ] 표기 보존: 학점·학기·비고 칸 원본값에서 대괄호(반각/전각) 탐지.
        # _to_num/_norm이 대괄호를 버리므로, 원본을 따로 스캔해 (칸이름, 원본값)으로 남긴다.
        _cols = {5: "기본학점", 6: "운영학점", 7: "1-1", 8: "1-2", 9: "2-1",
                 10: "2-2", 11: "3-1", 12: "3-2", 13: "비고"}
        교차이수표기 = []
        for c, lab in _cols.items():
            raw = cell(r, c)
            raw = "" if raw is None else str(raw).strip()
            if any(b in raw for b in ("[", "]", "［", "］")):
                교차이수표기.append((lab, raw))
        # 과목별 학기 배치 합(병합 그룹칸 제외) + 그룹배치 여부
        배치합 = 0.0
        그룹배치 = False
        for c in range(7, 13):  # 1-1 ~ 3-2
            if (r, c) in vmerged:
                if cell(r, c) is not None:
                    그룹배치 = True
                continue
            v = _to_num(cell(r, c))
            if v is not None:
                배치합 += v
        rows.append({
            "row": r,
            "교과군": gun,
            "과목유형": _norm_subject_type(typ),
            "과목명": subj,
            "기본학점": _to_num(cell(r, 5)),
            "운영학점": _to_num(cell(r, 6)),
            "1-1": _to_num(cell(r, 7)),
            "1-2": _to_num(cell(r, 8)),
            "2-1": _to_num(cell(r, 9)),
            "2-2": _to_num(cell(r, 10)),
            "3-1": _to_num(cell(r, 11)),
            "3-2": _to_num(cell(r, 12)),
            "비고": _norm(cell(r, 13)),
            "교차이수표기": 교차이수표기,
            "배치합": 배치합,
            "그룹배치": 그룹배치,
        })

    # ----- 합계 영역 파싱 -----
    # 필수 이수 학점 열을 헤더 라벨로 탐지한다(고정 열 번호를 쓰지 않는다).
    # 소계행에서 이 열을 읽어 A3(필수 이수 ≥ 84)를 자동 판정할 수 있다.
    req_col = None
    for c in range(1, ws.max_column + 1):
        for hr in range(h, h + 3):
            if "필수이수학점" in _norm(cell(hr, c)).replace(" ", ""):
                req_col = c
                break
        if req_col:
            break

    summary = {}
    for r in range(ss, min(ws.max_row, ss + 10) + 1):
        for c in range(1, 5):
            lbl = _norm(cell(r, c))
            if not lbl:
                continue
            terms = [_to_num(cell(r, c2)) for c2 in range(7, 13)]
            if "교과 이수 학점 소계" in lbl:
                summary["교과소계"] = _to_num(cell(r, 6))
                summary["교과학기"] = terms
                if req_col:
                    summary["필수이수"] = _to_num(cell(r, req_col))
            if "창의적 체험활동" in lbl:
                summary["창체"] = _to_num(cell(r, 6))
                summary["창체학기"] = terms
            if ("총 이수" in lbl or "학년별" in lbl) and "총학기" not in summary:
                summary["총학기"] = terms

    # ----- school 추출 (타입 분기) -----
    if isinstance(source, (str, Path)):
        school = Path(source).stem.split("_")[0]
    else:
        # BytesIO 등 파일 객체 - 호출자가 별도로 school을 채워야 함
        school = "업로드 파일"

    return {
        "school": school,
        "sheet": sheet,
        "header_row": h,
        "data_start": ds,
        "summary_start": ss,
        "rows": rows,
        "summary": summary,
        "정규과목": 정규과목,
    }
